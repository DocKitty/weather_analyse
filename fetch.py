from bs4 import BeautifulSoup
from bs4.element import Tag
import requests
import re
import openpyxl
import os
import logging
import datetime

url_base = 'https://www.tianqihoubao.com/lishi/'
logger = logging.getLogger(__name__)

def fetch_weather_data(file: str | openpyxl.Workbook, city: str, year: str, month: str) -> None:
    '''Fetch the weather data of a specific city, year and month from TianqiHouBao website and save it to a .xlsl file

    Args:
        file (str | openpyxl.Workbook): The file to save the data, it can be a file name or an openpyxl.Workbook object.\n
        For Workbook, it WILL NOT be saved automatically, you should save it when you need to.
        city (str): The city to fetch data.
        year (str): The year to fetch data, must be in the range of 2011 to current year.
        month (str): The month to fetch data.
    
    Returns:
        None
    '''
    if int(year) < 2011 or int(year) > datetime.datetime.now().year:
        logger.error(f'Year {year} is out of range')
        return

    if int(month) < 1 or int(month) > 12 or (int(year) == datetime.datetime.now().year and int(month) > datetime.datetime.now().month):
        logger.error(f'Month {month} is out of range')
        return
    
    if len(month) == 1:
        month = '0' + month

    file_path = ''
    if isinstance(file, str):
        file_path = os.path.join('data', file + '.xlsx')
        logger.info(f'Fetching data from TianqiHouBao website and saving to {file_path}...')
        if not os.path.exists(file_path):
            os.makedirs('data', exist_ok=True)
            wb = openpyxl.Workbook()
            sheet = wb.active
            logger.info(f'{file_path} does not exist, creating it...')
            if sheet == None:
                logger.error('Failed to create sheet')
                return
            sheet.append(['年', '月', '日', '白天天气', '夜间天气', '最高温度/℃', '最低温度/℃', '白天风力风向', '夜间风力风向'])
            wb.save(file_path)
            logger.info(f'{file_path} created')
        else:
            wb = openpyxl.load_workbook(file_path)
            logger.info(f'Found {file_path}, loading it...')
            sheet = wb.active
            if sheet == None:
                logger.error('Failed to load sheet')
                return
            logger.info(f'{file_path} loaded, new data will be appended to it')
    else:
        logger.info('Fetching data from TianqiHouBao website and saving to memory...')
        wb = file
        sheet = wb.active
        if sheet == None:
            logger.error('Failed to load sheet')
            return
        logger.info('Sheet loaded')

    logger.info(f'Fetching weather data for {year}-{month}...')

    url = f'{url_base}{city}/month/{year}{month}.html'
    response = requests.get(url)

    if response.status_code == 200:
        logger.info('Data fetched successfully')
        soup = BeautifulSoup(response.text, 'lxml')
        table = soup.find(class_='weather-table')
        logger.info('Finding weather data table...')
        if isinstance(table, Tag):
            rows = table.find_all('tr')
            logger.info(f'Found weather data table for {year}-{month}')
            for i in range(1, len(rows)): # skip table head
                row = rows[i]
                logger.info(f'Processing row: {i}')
                if isinstance(row, Tag):
                    columns = row.find_all('td')

                    dates = re.match('(.*)年(.*)月(.*)日', columns[0].text)
                    if (dates == None) or (len(dates.groups()) != 3):
                        logger.info('Date data not match, skip this row')
                        continue
                    else:
                        day : str = dates.group(3).lstrip('0')
                        logger.info(f'Found weather data of {year}-{month}-{day.zfill(2)}, processing...')

                    weathers : list[str] = columns[1].text.split('/')
                    if len(weathers) == 2:
                        weather_day = weathers[0].strip()
                        weather_night = weathers[1].strip()
                    else:
                        weather_day = ''
                        weather_night = ''
                        logger.info('Weather data not match pattern, skipping weather data')

                    temperatures : list[str] = columns[2].text.split('/')
                    if len(temperatures) == 2:
                        temperature_high = temperatures[0].strip().rstrip('℃')
                        temperature_low = temperatures[1].strip().rstrip('℃')
                    else:
                        temperature_high = ''
                        temperature_low = ''
                        logger.info('Temperature data not match pattern, skipping temperature data')

                    winds : list[str] = columns[3].text.split('/')
                    if len(winds) == 2:
                        wind_day = winds[0].strip()
                        wind_night = winds[1].strip()
                    else:
                        wind_day = ''
                        wind_night = ''
                        logger.info('Wind data not match pattern, skipping wind data')
                    
                    sheet.append([year, month.lstrip('0'), day, weather_day, weather_night, temperature_high, temperature_low, wind_day, wind_night])
                    logger.info(f'Fetched data for {year}-{month}-{day.zfill(2)}: {weather_day}, {weather_night}, {temperature_high}, {temperature_low}, {wind_day}, {wind_night}')
                else:
                    logger.error(f'Failed to process row: {row}, not a tag element')
            if isinstance(file, str):
                logger.info(f'Finished processing weather data for {year}-{month}, saving to {file_path}')
                wb.save(file_path)
            logger.info(f'Finished processing weather data for {year}-{month}, saved in memory')
        else:
            logger.error('Weather data table not found')
    else:
        logger.error('Unable to access weather data page')

def fetch_weather_data_year(file: str, city: str, year: str) -> None:
    '''Fetch the weather data of a specific city and year from TianqiHouBao website and save it to a .xlsl file

    Args:
        file_name (str): The file name to save the data
        city (str): The city to fetch data
        year (str): The year to fetch data
    
    Returns:
        None
    '''
    file_path = os.path.join('data', file + '.xlsx')
    if not os.path.exists(file_path):
            os.makedirs('data', exist_ok=True)
            wb = openpyxl.Workbook()
            sheet = wb.active
            logger.info(f'{file_path} does not exist, creating new file...')
            if sheet == None:
                logger.error('Failed to create sheet')
                return
            sheet.append(['年', '月', '日', '白天天气', '夜间天气', '最高温度/℃', '最低温度/℃', '白天风力风向', '夜间风力风向'])
            wb.save(file_path)
            logger.info(f'{file_path} created')
    else:
        wb = openpyxl.load_workbook(file_path)
        logger.info(f'Found existing {file_path}, loading...')

    logger.info(f'Fetching weather data for {city} in {year}...')
    for month in range(1, 13):
        fetch_weather_data(wb, city, year, str(month))

    wb.save(file_path)
    wb.close()
    logger.info(f'Weather data fetched and saved to {file_path}')